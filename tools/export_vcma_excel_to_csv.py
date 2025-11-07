import argparse
import os
import re
from datetime import datetime, timedelta
from typing import Optional, Tuple

import pandas as pd
from openpyxl import load_workbook


def detect_header_and_seed(excel_path: str, sheet_index: int = 0) -> Tuple[int, Optional[datetime]]:
    """
    Scan the first ~50 rows to find:
    - The header row index (0-based overall, for pandas read_excel header=...)
    - A seed datetime from metadata (Filter Applied: Before ..., or Date/Time of Report)
    """
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.worksheets[sheet_index]

    header_row_idx = None
    seed_dt: Optional[datetime] = None
    report_dt: Optional[datetime] = None

    rows_to_scan = 60
    for ridx, row in enumerate(ws.iter_rows(min_row=1, max_row=rows_to_scan, values_only=True), start=1):
        # Normalize string values in row
        vals = [str(v).strip() if v is not None else "" for v in row]
        joined = ",".join(vals)

        # Parse metadata
        if not seed_dt:
            m = re.search(r"Filter\s*Applied\s*:\s*Event\s*Time\s*:\s*Before\s*(.+)$", joined, re.IGNORECASE)
            if m:
                try:
                    seed_dt = pd.to_datetime(m.group(1), errors="coerce")
                    if pd.isna(seed_dt):
                        seed_dt = None
                except Exception:
                    seed_dt = None
        if not report_dt:
            m2 = re.search(r"Date/Time\s*of\s*Report\s*:\s*(.+)$", joined, re.IGNORECASE)
            if m2:
                try:
                    report_dt = pd.to_datetime(m2.group(1), errors="coerce")
                    if pd.isna(report_dt):
                        report_dt = None
                except Exception:
                    report_dt = None

        # Detect header row: must include at least Event Time + Source
        if header_row_idx is None:
            has_event_time = any(v.strip().lower() == "event time" for v in vals)
            has_source = any(v.strip().lower() == "source" for v in vals)
            if has_event_time and has_source:
                # openpyxl is 1-based, pandas header wants 0-based
                header_row_idx = ridx - 1
                # No need to scan more for header

    # Prefer filter date as seed, else report date
    seed_dt_final = seed_dt or report_dt
    if isinstance(seed_dt_final, pd.Timestamp):
        seed_dt_final = seed_dt_final.to_pydatetime()
    return header_row_idx if header_row_idx is not None else 0, seed_dt_final


def reconstruct_event_time(df: pd.DataFrame, seed_dt: Optional[datetime]) -> pd.Series:
    """
    Reconstruct full datetimes for the 'Event Time' column using logic compatible with
    the backend loader:
      - Full datetime strings parsed as-is
      - Time-only (HH:MM:SS [AM/PM]) get date via forward-fill
      - Truncated MM:SS(.fff) rows update the last known full timestamp's minute/second/us
      - If no prior full timestamp is available, use seed_dt
    Returns a pandas Series[datetime64[ns]].
    """
    if "Event Time" not in df.columns:
        raise ValueError("Column 'Event Time' not found in sheet")

    et_raw = df["Event Time"]

    # Normalize all to string for pattern detection, keep a parallel object list for types
    et_str = et_raw.astype(str).str.strip()
    et_str = et_str.str.replace(r"\s+", " ", regex=True)

    # Regex patterns
    pattern_full_ampm = r"^\d{1,2}/\d{1,2}/\d{4}\s+\d{1,2}:\d{2}:\d{2}\s+(AM|PM)$"
    pattern_full_iso = r"^\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2}$"
    pattern_time_ampm = r"^\d{1,2}:\d{2}:\d{2}\s+(AM|PM)$"
    pattern_time_24h = r"^\d{1,2}:\d{2}:\d{2}$"
    pattern_mmss = r"^\d{1,2}:\d{2}(\.\d+)?$"

    # Additional patterns with milliseconds and HH:MM (no seconds)
    pattern_full_ampm_ms = r"^\d{1,2}/\d{1,2}/\d{4}\s+\d{1,2}:\d{2}:\d{2}\.\d+\s+(AM|PM)$"
    pattern_full_iso_ms = r"^\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2}\.\d+$"
    pattern_time_ampm_ms = r"^\d{1,2}:\d{2}:\d{2}\.\d+\s+(AM|PM)$"
    pattern_time_24h_ms = r"^\d{1,2}:\d{2}:\d{2}\.\d+$"
    pattern_time_ampm_hm = r"^\d{1,2}:\d{2}\s+(AM|PM)$"
    pattern_time_24h_hm = r"^\d{1,2}:\d{2}$"

    is_full_ampm = et_str.str.match(pattern_full_ampm, na=False)
    is_full_iso = et_str.str.match(pattern_full_iso, na=False)
    is_time_ampm = et_str.str.match(pattern_time_ampm, na=False)
    is_time_24h = et_str.str.match(pattern_time_24h, na=False)
    is_mmss = et_str.str.match(pattern_mmss, na=False)
    is_full_ampm_ms = et_str.str.match(pattern_full_ampm_ms, na=False)
    is_full_iso_ms = et_str.str.match(pattern_full_iso_ms, na=False)
    is_time_ampm_ms = et_str.str.match(pattern_time_ampm_ms, na=False)
    is_time_24h_ms = et_str.str.match(pattern_time_24h_ms, na=False)
    is_time_ampm_hm = et_str.str.match(pattern_time_ampm_hm, na=False)
    # Careful: HH:MM (no seconds) overlaps with MM:SS pattern. We'll treat HH:MM only when not matching MM:SS.
    is_time_24h_hm = et_str.str.match(pattern_time_24h_hm, na=False) & (~is_mmss)

    # Disambiguate HH:MM:SS vs MM:SS
    ambiguous = is_time_24h & is_mmss
    if ambiguous.any():
        colon_count = et_str[ambiguous].str.count(":")
        is_time_24h.loc[ambiguous] = (colon_count == 2)
        is_mmss.loc[ambiguous] = (colon_count == 1)

    # Prepare result Series
    parsed = pd.Series([pd.NaT] * len(df), index=df.index, dtype="datetime64[ns]")

    # 1) Full datetime rows
    if is_full_ampm_ms.any():
        parsed.loc[is_full_ampm_ms] = pd.to_datetime(et_str[is_full_ampm_ms], format="%m/%d/%Y %I:%M:%S.%f %p", errors="coerce")
    if is_full_iso_ms.any():
        parsed.loc[is_full_iso_ms] = pd.to_datetime(et_str[is_full_iso_ms], format="%Y-%m-%d %H:%M:%S.%f", errors="coerce")
    if is_full_ampm.any():
        parsed.loc[is_full_ampm] = pd.to_datetime(et_str[is_full_ampm], format="%m/%d/%Y %I:%M:%S %p", errors="coerce")
    if is_full_iso.any():
        parsed.loc[is_full_iso] = pd.to_datetime(et_str[is_full_iso], format="%Y-%m-%d %H:%M:%S", errors="coerce")

    # 2) If some cells are already true datetimes (Excel), include them
    # Pandas may import Excel datetimes as datetime64 directly; fill where parsed is NaT
    dt_mask = et_raw.apply(lambda v: isinstance(v, (pd.Timestamp, datetime)))
    if dt_mask.any():
        parsed.loc[dt_mask & parsed.isna()] = pd.to_datetime(et_raw[dt_mask], errors="coerce")

    # Forward-fill full timestamps and dates for reconstruction
    ff_full_ts = parsed.ffill()

    # If there's no full timestamp at all and a seed is present, seed the first row
    if ff_full_ts.isna().all() and seed_dt is not None:
        ff_full_ts.iloc[0] = pd.Timestamp(seed_dt)
        ff_full_ts = ff_full_ts.ffill()

    # 3) Time-only rows (AM/PM)
    time_only_ampm = is_time_ampm | is_time_ampm_ms | is_time_ampm_hm
    if time_only_ampm.any():
        # Build combined strings from forward-filled date + time
        base_dates = ff_full_ts.dt.date
        mask_common = base_dates.notna()
        # AM/PM with milliseconds
        mask_ms = is_time_ampm_ms & mask_common
        if mask_ms.any():
            combined_ms = base_dates.astype(str) + " " + et_str
            parsed.loc[mask_ms] = pd.to_datetime(combined_ms[mask_ms], format="%Y-%m-%d %I:%M:%S.%f %p", errors="coerce")
        # AM/PM HH:MM:SS
        mask_hms = is_time_ampm & mask_common
        if mask_hms.any():
            combined_hms = base_dates.astype(str) + " " + et_str
            parsed.loc[mask_hms] = pd.to_datetime(combined_hms[mask_hms], format="%Y-%m-%d %I:%M:%S %p", errors="coerce")
        # AM/PM HH:MM
        mask_hm = is_time_ampm_hm & mask_common
        if mask_hm.any():
            combined_hm = base_dates.astype(str) + " " + et_str
            parsed.loc[mask_hm] = pd.to_datetime(combined_hm[mask_hm], format="%Y-%m-%d %I:%M %p", errors="coerce")

    # 4) Time-only rows (24h)
    time_only_24 = is_time_24h | is_time_24h_ms | is_time_24h_hm
    if time_only_24.any():
        base_dates = ff_full_ts.dt.date
        mask_common = base_dates.notna()
        # 24h with milliseconds
        mask_ms = is_time_24h_ms & mask_common
        if mask_ms.any():
            combined_ms = base_dates.astype(str) + " " + et_str
            parsed.loc[mask_ms] = pd.to_datetime(combined_ms[mask_ms], format="%Y-%m-%d %H:%M:%S.%f", errors="coerce")
        # 24h HH:MM:SS
        mask_hms = is_time_24h & mask_common
        if mask_hms.any():
            combined_hms = base_dates.astype(str) + " " + et_str
            parsed.loc[mask_hms] = pd.to_datetime(combined_hms[mask_hms], format="%Y-%m-%d %H:%M:%S", errors="coerce")
        # 24h HH:MM (no seconds)
        mask_hm = is_time_24h_hm & mask_common
        if mask_hm.any():
            combined_hm = base_dates.astype(str) + " " + et_str
            parsed.loc[mask_hm] = pd.to_datetime(combined_hm[mask_hm], format="%Y-%m-%d %H:%M", errors="coerce")

    # Refresh forward-fill after any new full timestamps
    ff_full_ts = parsed.ffill()
    if ff_full_ts.isna().all() and seed_dt is not None:
        ff_full_ts.iloc[0] = pd.Timestamp(seed_dt)
        ff_full_ts = ff_full_ts.ffill()

    # 5) Truncated MM:SS(.fff) rows
    mmss_mask = is_mmss
    if mmss_mask.any():
        base_ts = ff_full_ts.copy()
        for idx in et_str[mmss_mask].index:
            base = base_ts.loc[idx]
            if pd.isna(base):
                # try seed
                if seed_dt is None:
                    continue
                base = pd.Timestamp(seed_dt)
            s = et_str.loc[idx]
            # Parse minutes, seconds, microseconds
            try:
                parts = s.split(":")
                minutes = int(parts[0])
                sec_parts = parts[1].split(".")
                seconds = int(sec_parts[0])
                micro = 0
                if len(sec_parts) > 1:
                    frac = sec_parts[1].ljust(6, "0")[:6]
                    micro = int(frac)
                # Replace minute/second/microsecond on base
                new_ts = base.replace(minute=minutes, second=seconds, microsecond=micro)
                parsed.loc[idx] = new_ts
            except Exception:
                # leave as NaT
                pass

    return parsed


def main():
    ap = argparse.ArgumentParser(description="Export VCMA Excel to CSV with full Event Time reconstruction.")
    ap.add_argument("--excel", required=True, help="Path to source Excel (.xlsm/.xlsx)")
    ap.add_argument("--sheet-index", type=int, default=0, help="Sheet index (0-based). Default: 0 (first sheet)")
    ap.add_argument("--out", default=None, help="Output CSV path. Default: <excel_dir>/VCMA_clean.csv")
    args = ap.parse_args()

    excel_path = os.path.abspath(args.excel)
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    # Detect header & seed
    header_row, seed_dt = detect_header_and_seed(excel_path, sheet_index=args.sheet_index)

    # Read with pandas using detected header
    df = pd.read_excel(excel_path, sheet_name=args.sheet_index, header=header_row, engine="openpyxl")
    if df.empty:
        raise ValueError("No data rows found after header detection")

    # Reconstruct Event Time
    parsed_dt = reconstruct_event_time(df, seed_dt)

    # Keep ALL rows. Format parsed datetimes; fallback to original text for unparsed
    total_rows = len(df)
    parsed_mask = parsed_dt.notna()
    # Start with original values as strings
    event_time_out = df["Event Time"].astype(str).str.strip()
    # Fill where parsed
    if parsed_mask.any():
        formatted = parsed_dt.loc[parsed_mask].dt.strftime("%m/%d/%Y %I:%M:%S.%f %p")
        # Trim microseconds to milliseconds for readability
        formatted = formatted.str.replace(r"(\.\d{3})\d{3}", r"\1", regex=True)
        event_time_out.loc[parsed_mask] = formatted
    df["Event Time"] = event_time_out

    # Output path
    out_path = args.out or os.path.join(os.path.dirname(excel_path), "VCMA_clean.csv")

    # Write CSV
    df.to_csv(out_path, index=False, encoding="utf-8")
    parsed_count = int(parsed_mask.sum())
    unparsed_count = int(total_rows - parsed_count)
    print(f"✓ Wrote cleaned CSV with full datetimes: {out_path}")
    print(f"Rows: {len(df)} (from {total_rows})  Columns: {len(df.columns)}")
    print(f"Parsed Event Time: {parsed_count}  Unparsed (kept raw): {unparsed_count}")


if __name__ == "__main__":
    main()
